"""Tests for Excel worksheet generation."""

import json
from pathlib import Path
from io import BytesIO
import pytest
from openpyxl import load_workbook
from recon.models import EventOrder
from recon.builder import compute_totals, generate_excel


@pytest.fixture
def beo_2895() -> EventOrder:
    """Load the golden fixture."""
    fixture_path = Path(__file__).parent / "fixtures" / "beo_2895.json"
    data = json.loads(fixture_path.read_text())
    return EventOrder.model_validate(data)


class TestExcelExport:
    def test_generates_valid_excel(self, beo_2895: EventOrder):
        """generate_excel returns a valid Excel file."""
        worksheet_output = compute_totals(beo_2895)
        excel_bytes = generate_excel(worksheet_output)

        # Should be readable as Excel
        wb = load_workbook(BytesIO(excel_bytes))
        assert "Totals" in wb.sheetnames
        assert "Line Items" in wb.sheetnames

    def test_totals_sheet_has_correct_values(self, beo_2895: EventOrder):
        """Totals sheet contains the correct category totals."""
        worksheet_output = compute_totals(beo_2895)
        excel_bytes = generate_excel(worksheet_output)

        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["Totals"]

        # Header row
        assert ws["A1"].value == "Category"
        assert ws["B1"].value == "Delphi (incl cash)"
        assert ws["C1"].value == "Opera (excl cash)"

        # Find Food row and check values
        for row in range(2, ws.max_row + 1):
            if ws[f"A{row}"].value == "food":
                assert ws[f"B{row}"].value == 123466.00
                assert ws[f"C{row}"].value == 123466.00
                break

    def test_line_items_sheet_has_all_items(self, beo_2895: EventOrder):
        """Line Items sheet contains all line items."""
        worksheet_output = compute_totals(beo_2895)
        excel_bytes = generate_excel(worksheet_output)

        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["Line Items"]

        # Should have header + 12 line items
        assert ws.max_row == 13  # 1 header + 12 items
