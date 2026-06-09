"""Tests for Delphi posting report adapter."""

from io import BytesIO
import pytest
from openpyxl import Workbook
from recon.delphi_adapter import parse_delphi_report


def create_sample_delphi_excel() -> bytes:
    """Create a sample Delphi export for testing."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Posting Report"

    # Header row
    ws["A1"] = "Category"
    ws["B1"] = "Amount"

    # Data rows
    data = [
        ("Food", 123466.00),
        ("Beverage", 81839.60),
        ("Resource", 1150.00),
        ("Other", 3124.00),
        ("Venue Hire", 500.00),
    ]
    for row_idx, (category, amount) in enumerate(data, 2):
        ws[f"A{row_idx}"] = category
        ws[f"B{row_idx}"] = amount

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


class TestDelphiAdapter:
    def test_parse_delphi_report(self):
        """Parse Delphi Excel and extract category totals."""
        excel_bytes = create_sample_delphi_excel()
        result = parse_delphi_report(BytesIO(excel_bytes))

        assert result["food"] == 123466.00
        assert result["beverage"] == 81839.60
        assert result["resource"] == 1150.00
        assert result["other"] == 3124.00
        assert result["venue_hire"] == 500.00

    def test_normalizes_category_names(self):
        """Category names are normalized to lowercase with underscores."""
        excel_bytes = create_sample_delphi_excel()
        result = parse_delphi_report(BytesIO(excel_bytes))

        # "Venue Hire" should become "venue_hire"
        assert "venue_hire" in result
        assert "Venue Hire" not in result
