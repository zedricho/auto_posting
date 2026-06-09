"""Adapter for parsing Delphi posting reports."""

from io import BytesIO
from typing import BinaryIO, Dict
from openpyxl import load_workbook


# Mapping from Delphi category names to our normalized names
CATEGORY_MAPPING = {
    "food": "food",
    "beverage": "beverage",
    "resource": "resource",
    "other": "other",
    "venue hire": "venue_hire",
    "venue_hire": "venue_hire",
    "av": "av",
    "audio visual": "av",
}


def _normalize_category(name: str) -> str:
    """Normalize a category name to our standard format."""
    normalized = name.lower().strip()
    return CATEGORY_MAPPING.get(normalized, normalized.replace(" ", "_"))


def parse_delphi_report(file: BinaryIO) -> Dict[str, float]:
    """
    Parse a Delphi posting report Excel file.

    Expects columns: Category, Amount (or similar).
    Returns dict mapping normalized category names to amounts.
    """
    wb = load_workbook(file, data_only=True)
    ws = wb.active

    # Find the header row and identify columns
    category_col = None
    amount_col = None

    for col in range(1, ws.max_column + 1):
        header = str(ws.cell(row=1, column=col).value or "").lower()
        if "category" in header or "account" in header:
            category_col = col
        elif "amount" in header or "total" in header:
            amount_col = col

    if category_col is None or amount_col is None:
        raise ValueError("Could not find Category and Amount columns in Delphi report")

    # Parse data rows
    result: dict[str, float] = {}
    for row in range(2, ws.max_row + 1):
        category_cell = ws.cell(row=row, column=category_col).value
        amount_cell = ws.cell(row=row, column=amount_col).value

        if category_cell is None:
            continue

        category = _normalize_category(str(category_cell))
        amount = float(amount_cell) if amount_cell is not None else 0.0

        # Accumulate in case there are multiple rows per category
        result[category] = result.get(category, 0.0) + amount

    return result
