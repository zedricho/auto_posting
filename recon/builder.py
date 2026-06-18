"""Builder: compute category totals and generate worksheet."""

from collections import defaultdict
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from recon.models import EventOrder, CategoryTotals, WorksheetOutput


def compute_totals(event: EventOrder) -> WorksheetOutput:
    """
    Compute Delphi and Opera totals from an EventOrder.

    - Delphi total = all lines (contracted + consumption + cash)
    - Opera total = exclude cash (only contracted + consumption)
    - Venue hire (shortfall) is reduced by consumption amounts
    """
    # Calculate consumption total (to reduce venue hire shortfall)
    consumption_total = sum(
        item.value for item in event.line_items
        if item.money_type == "consumption" and item.value > 0
    )

    # Group line items by category
    by_category: dict[str, list] = defaultdict(list)
    for item in event.line_items:
        by_category[item.category].append(item)

    totals = []
    for category, items in by_category.items():
        # Delphi = all lines
        delphi_total = sum(item.value for item in items)

        # Opera = exclude cash (money_type != "cash")
        opera_total = sum(
            item.value for item in items if item.money_type != "cash"
        )

        # Venue hire (shortfall) is reduced by consumption revenue
        # Consumption adds to F&B, which reduces the minimum spend shortfall
        if category == "venue_hire" and consumption_total > 0:
            reduction = min(delphi_total, consumption_total)
            delphi_total = max(0, delphi_total - reduction)
            opera_total = max(0, opera_total - reduction)

        totals.append(
            CategoryTotals(
                category=category,
                delphi_total=round(delphi_total, 2),
                opera_total=round(opera_total, 2),
            )
        )

    # Sort totals by category for consistent output
    category_order = ["food", "beverage", "resource", "other", "venue_hire", "av"]
    totals.sort(key=lambda t: category_order.index(t.category) if t.category in category_order else 99)

    return WorksheetOutput(
        event=event,
        totals=totals,
        delphi_grand_total=round(sum(t.delphi_total for t in totals), 2),
        opera_grand_total=round(sum(t.opera_total for t in totals), 2),
    )


def generate_excel(output: WorksheetOutput) -> bytes:
    """
    Generate an Excel workbook from WorksheetOutput.

    Returns the workbook as bytes (ready for download).
    """
    wb = Workbook()

    # Sheet 1: Totals
    ws_totals = wb.active
    ws_totals.title = "Totals"

    # Header row
    headers = ["Category", "Delphi (incl cash)", "Opera (excl cash)"]
    for col, header in enumerate(headers, 1):
        cell = ws_totals.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    # Data rows
    for row_idx, total in enumerate(output.totals, 2):
        ws_totals.cell(row=row_idx, column=1, value=total.category)
        ws_totals.cell(row=row_idx, column=2, value=total.delphi_total)
        ws_totals.cell(row=row_idx, column=3, value=total.opera_total)

    # Grand total row
    grand_row = len(output.totals) + 2
    ws_totals.cell(row=grand_row, column=1, value="TOTAL").font = Font(bold=True)
    ws_totals.cell(row=grand_row, column=2, value=output.delphi_grand_total).font = Font(bold=True)
    ws_totals.cell(row=grand_row, column=3, value=output.opera_grand_total).font = Font(bold=True)

    # Column widths
    ws_totals.column_dimensions["A"].width = 15
    ws_totals.column_dimensions["B"].width = 20
    ws_totals.column_dimensions["C"].width = 20

    # Sheet 2: Line Items
    ws_items = wb.create_sheet("Line Items")

    item_headers = ["Category", "Type", "Basis", "Qty/Pax", "Unit Price", "Value", "Money Type", "Posts To"]
    for col, header in enumerate(item_headers, 1):
        cell = ws_items.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    for row_idx, item in enumerate(output.event.line_items, 2):
        ws_items.cell(row=row_idx, column=1, value=item.category)
        ws_items.cell(row=row_idx, column=2, value=item.type)
        ws_items.cell(row=row_idx, column=3, value=item.basis)
        ws_items.cell(row=row_idx, column=4, value=item.pax or item.qty or item.guards or "")
        ws_items.cell(row=row_idx, column=5, value=item.unit_price or "")
        ws_items.cell(row=row_idx, column=6, value=item.value)
        ws_items.cell(row=row_idx, column=7, value=item.money_type)
        ws_items.cell(row=row_idx, column=8, value=item.posts_to)

    # Column widths for line items
    for col in range(1, 9):
        ws_items.column_dimensions[get_column_letter(col)].width = 18

    # Save to bytes
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
